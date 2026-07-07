import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Print all cells with their values and fill colors
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        fg = fill.fgColor
        bg = fill.bgColor
        
        # Get color info
        fg_color = None
        bg_color = None
        fill_type = fill.fill_type
        
        if fg and fg.type == 'rgb':
            fg_color = fg.rgb
        elif fg and fg.type == 'theme':
            fg_color = f"theme:{fg.theme}"
        elif fg and fg.type == 'indexed':
            fg_color = f"indexed:{fg.indexed}"
            
        if bg and bg.type == 'rgb':
            bg_color = bg.rgb
        
        val = cell.value
        if val is not None or (fg_color and fg_color != '00000000' and fill_type):
            col_letter = get_column_letter(col)
            print(f"  {col_letter}{row}: value={repr(val)}, fill_type={fill_type}, fg={fg_color}, bg={bg_color}")
