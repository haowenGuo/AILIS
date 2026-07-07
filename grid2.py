import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb.active

# Let's print a very clean grid with exact colors
# Also check: does "move two cells" mean moving 2 steps where each step is 1 cell,
# possibly changing direction? Let's first make sure we have the grid right.

# Print as a grid using single characters
# . = blue/blocked, other chars for non-blue
for row in range(1, ws.max_row + 1):
    line = ""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        fill = cell.fill
        fg = fill.fgColor
        fg_color = None
        if fg and fg.type == 'rgb':
            fg_color = fg.rgb
        elif fg and fg.type == 'theme':
            fg_color = f"theme:{fg.theme}"
        
        if val == 'START':
            line += "S"
        elif val == 'END':
            line += "E"
        elif fg_color == 'FF0099FF':
            line += "."  # blue blocked
        elif fg_color == 'FF92D050':
            line += "G"  # green
        elif fg_color == 'FFF478A7':
            line += "P"  # pink
        elif fg_color == 'FFFFFF00':
            line += "Y"  # yellow
        elif fg_color and str(fg_color).startswith('theme'):
            line += "W"
        else:
            line += "?"
    print(f"{row:2d}|{line}|")

print("   ABCDEFGHI")
