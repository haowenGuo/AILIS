import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb.active

# Color mapping for display
# FF0099FF = blue (blocked)
# FF92D050 = green
# FFF478A7 = pink
# FFFFFF00 = yellow
# theme:0 = white/default (START/END)

color_short = {
    'FF0099FF': 'B',  # blue - blocked
    'FF92D050': 'G',  # green
    'FFF478A7': 'P',  # pink
    'FFFFFF00': 'Y',  # yellow
}

print("Grid (rows 1-20, cols A-I):")
print("    A  B  C  D  E  F  G  H  I")
for row in range(1, ws.max_row + 1):
    line = f"{row:2d}: "
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        fill = cell.fill
        fg = fill.fgColor
        fg_color = None
        if fg and fg.type == 'rgb':
            fg_color = fg.rgb
        elif fg and fg.type == 'theme':
            fg_color = f"theme:{fg.theme}"
        
        if val == 'START':
            line += "ST "
        elif val == 'END':
            line += "EN "
        elif fg_color in color_short:
            line += f"{color_short[fg_color]:2} "
        elif fg_color and fg_color.startswith('theme'):
            line += "W  "  # white/default
        else:
            line += "?  "
            print(f"  Unknown color at {cell.coordinate}: {fg_color}")
    print(line)

print()
print("Color hex reference:")
print("B (blue/blocked): FF0099FF -> 0099FF")
print("G (green): FF92D050 -> 92D050")
print("P (pink): FFF478A7 -> F478A7")
print("Y (yellow): FFFFFF00 -> FFFF00")
print("W (white/theme0): FFFFFFFF -> FFFFFF")
