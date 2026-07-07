import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

# Build a grid of colors and values
# Color codes:
# FF0099FF = blue (obstacle)
# FF92D050 = green
# FFFFFF00 = yellow
# FFF478A7 = pink
# theme:0 = white/black text (START/END)

rows = ws.max_row
cols = ws.max_column

# Create a color grid (fg RGB, stripping the FF alpha prefix for display)
grid = {}
value_grid = {}
for r in range(1, rows+1):
    for c in range(1, cols+1):
        cell = ws.cell(row=r, column=c)
        fill = cell.fill
        fg = fill.fgColor
        color = None
        if fg.type == 'rgb':
            color = fg.rgb
        elif fg.type == 'theme':
            color = f"theme:{fg.theme}"
        elif fg.type == 'indexed':
            color = f"idx:{fg.indexed}"
        grid[(r,c)] = color
        value_grid[(r,c)] = cell.value

# Print visual map
# Legend: B=blue, G=green, Y=yellow, P=pink, S=START, E=END, W=white(theme:0)
def symbol(r, c):
    v = value_grid[(r,c)]
    if v == 'START': return 'S'
    if v == 'END': return 'E'
    col = grid[(r,c)]
    if col == 'FF0099FF': return 'B'
    if col == 'FF92D050': return 'G'
    if col == 'FFFFFF00': return 'Y'
    if col == 'FFF478A7': return 'P'
    if col and col.startswith('theme'): return 'W'
    return '?'

print("Map (rows 1-20, cols A-I = 1-9):")
print("    A B C D E F G H I")
for r in range(1, rows+1):
    row_str = f" {r:2d} "
    for c in range(1, cols+1):
        row_str += symbol(r,c) + ' '
    print(row_str)

print()
print("Color key:")
print("  B = FF0099FF (blue - obstacle)")
print("  G = FF92D050 (green)")
print("  Y = FFFFFF00 (yellow)")
print("  P = FFF478A7 (pink)")
print("  S = START (theme:0)")
print("  E = END (theme:0)")
print("  W = theme:0 (white/default)")

# Also print the full color grid for reference
print()
print("Full color grid (6-digit hex without alpha):")
for r in range(1, rows+1):
    for c in range(1, cols+1):
        col = grid[(r,c)]
        v = value_grid[(r,c)]
        col_letter = get_column_letter(c)
        hex6 = col[2:] if col and col.startswith('FF') and len(col)==8 else col
        if v:
            print(f"  {col_letter}{r}: {hex6} ({v})")
        else:
            print(f"  {col_letter}{r}: {hex6}")
