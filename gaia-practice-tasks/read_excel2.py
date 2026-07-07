import openpyxl
from openpyxl.styles import PatternFill, Color

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx', data_only=False)
ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()

# Let's also print the merged cells
print(f'Merged cells: {ws.merged_cells.ranges}')
print()

# Print ALL cells info
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        val = cell.value
        fg = fill.fgColor
        bg = fill.bgColor
        
        fg_rgb = fg.rgb if fg else None
        bg_rgb = bg.rgb if bg else None
        fg_theme = fg.theme if fg else None
        fg_tint = fg.tint if fg else None
        fg_type = fg.type if fg else None
        fg_indexed = fg.indexed if fg else None
        
        if val is not None or (fg_rgb and fg_rgb not in ('00000000', '0')):
            print(f'Cell {cell.coordinate}: val={val!r}, fg_rgb={fg_rgb!r}, fg_theme={fg_theme!r}, fg_tint={fg_tint!r}, fg_type={fg_type!r}, fg_indexed={fg_indexed!r}')

print()
print("--- Now checking with pattern_type ---")
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        if fill.patternType:
            print(f'Cell {cell.coordinate}: patternType={fill.patternType}, fgColor.rgb={fill.fgColor.rgb!r}, bgColor.rgb={fill.bgColor.rgb!r}')
