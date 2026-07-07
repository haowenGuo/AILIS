import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx', data_only=False)
ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()

# Print all cells with their values and fill colors
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        fill = cell.fill
        fg_color = fill.fgColor
        color_str = None
        if fg_color:
            color_str = fg_color.rgb
        if color_str and color_str not in ('00000000', '0'):
            print(f'Cell {cell.coordinate}: value={cell.value!r}, fill_color={color_str}')
        elif cell.value is not None:
            print(f'Cell {cell.coordinate}: value={cell.value!r} (no fill)')
