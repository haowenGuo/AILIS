import openpyxl
from collections import deque

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

rows = ws.max_row
cols = ws.max_column

# Build grid
passable = {}
colors = {}
start = None
end = None

for r in range(1, rows+1):
    for c in range(1, cols+1):
        cell = ws.cell(row=r, column=c)
        fill = cell.fill
        fg = fill.fgColor
        color = None
        if fg.type == 'rgb':
            color = fg.rgb
        elif fg.type == 'theme':
            color = f"theme:{fg.theme}"
        colors[(r,c)] = color
        
        if cell.value == 'START':
            start = (r, c)
            passable[(r,c)] = True
        elif cell.value == 'END':
            end = (r, c)
            passable[(r,c)] = True
        elif color == 'FF0099FF':
            passable[(r,c)] = False
        else:
            passable[(r,c)] = True

print(f"START: {start}, END: {end}")

# Directions: (dr, dc, name, opposite_name)
DIRS = {
    'U': (-1, 0),
    'D': (1, 0),
    'L': (0, -1),
    'R': (0, 1),
}
OPP = {'U': 'D', 'D': 'U', 'L': 'R', 'R': 'L'}

# Interpretation: Each turn = two 1-cell steps.
# After each step, you move to adjacent cell.
# You cannot step on blue cells.
# You cannot move backward = cannot reverse direction (180 degrees) from your last step.
# State = (row, col, last_step_direction)

# BFS from START
# Turn 0: at START, no last direction
# Each turn: take 2 steps, each step to adjacent cell, no reversing

def get_neighbors(r, c, last_dir):
    """Get valid (nr, nc, dir) from (r,c) given last direction (can't go opposite)."""
    result = []
    for dname, (dr, dc) in DIRS.items():
        if last_dir is not None and dname == OPP[last_dir]:
            continue
        nr, nc = r + dr, c + dc
        if 1 <= nr <= rows and 1 <= nc <= cols and passable.get((nr, nc), False):
            result.append((nr, nc, dname))
    return result

# State after 0 turns: at start, no last direction
# After turn 1: took 2 steps, last direction is direction of step 2
# We need to track position after each step to enforce no-reverse within a turn

# Let's do BFS where each state is (r, c, last_dir) and we track turn number
# Turn 0: (start_r, start_c, None)
# To complete a turn, we take 2 steps:
#   Step 1: from (r,c,last_dir) go to (r1,c1,d1) where d1 != opposite(last_dir)
#   Step 2: from (r1,c1,d1) go to (r2,c2,d2) where d2 != opposite(d1)
# After turn: state is (r2, c2, d2)

states = {(start[0], start[1], None)}
all_paths = {}  # state -> path to get there (for debugging)
all_paths[(start[0], start[1], None)] = []

for turn in range(1, 12):
    next_states = {}
    for (r, c, last_dir) in states:
        # Step 1
        for (r1, c1, d1) in get_neighbors(r, c, last_dir):
            # Step 2
            for (r2, c2, d2) in get_neighbors(r1, c1, d1):
                new_state = (r2, c2, d2)
                if new_state not in next_states:
                    path = all_paths.get((r, c, last_dir), []) + [d1, d2]
                    next_states[new_state] = path
    states = set(next_states.keys())
    all_paths = next_states
    
    positions = set((s[0], s[1]) for s in states)
    print(f"\nTurn {turn}: {len(positions)} unique positions, {len(states)} states")
    for (r, c) in sorted(positions):
        col_letter = chr(ord('A') + c - 1)
        color = colors[(r,c)]
        hex6 = color[2:] if color and color.startswith('FF') and len(color)==8 else color
        print(f"  {col_letter}{r} (color={hex6})")

print(f"\n=== After turn 11 ===")
for (r, c, d) in sorted(states):
    col_letter = chr(ord('A') + c - 1)
    color = colors[(r,c)]
    hex6 = color[2:] if color and color.startswith('FF') and len(color)==8 else color
    path = all_paths[(r, c, d)]
    path_str = ''.join(path)
    print(f"  {col_letter}{r}, last_dir={d}, color={hex6}, path={path_str}")
