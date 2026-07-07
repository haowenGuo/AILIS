import openpyxl

wb = openpyxl.load_workbook("F:\\AILIS_self_evolution_runtime\\gaia-practice-tasks\\task2-excel-map.xlsx")
ws = wb.active

# Let me examine the fill more carefully, especially for START and END cells
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        val = cell.value
        
        fg_rgb = "N/A"
        bg_rgb = "N/A"
        pat = "N/A"
        try:
            if fill and fill.fgColor:
                fg_rgb = fill.fgColor.rgb
        except Exception as e:
            fg_rgb = f"ERROR: {e}"
        try:
            if fill and fill.bgColor:
                bg_rgb = fill.bgColor.rgb
        except Exception as e:
            bg_rgb = f"ERROR: {e}"
        try:
            if fill:
                pat = fill.patternType
        except Exception as e:
            pat = f"ERROR: {e}"
        
        # Print all non-empty cells or cells that have a fill
        if val is not None or pat not in [None, 'none', 'N/A']:
            print(f"{cell.coordinate}: val={repr(val)}, pat={pat}, fg={fg_rgb}, bg={bg_rgb}")
