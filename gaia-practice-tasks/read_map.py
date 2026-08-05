import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Print all cell values and their fill colors
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        fg = fill.fgColor
        bg = fill.bgColor
        
        # Get color info
        fg_color = None
        bg_color = None
        fill_type = fill.fill_type
        
        if fg and fg.type == 'rgb':
            fg_color = fg.rgb
        elif fg and fg.type == 'theme':
            fg_color = f"theme:{fg.theme}"
        elif fg and fg.type == 'indexed':
            fg_color = f"indexed:{fg.indexed}"
            
        if bg and bg.type == 'rgb':
            bg_color = bg.rgb
        elif bg and bg.type == 'theme':
            bg_color = f"theme:{bg.theme}"
        elif bg and bg.type == 'indexed':
            bg_color = f"indexed:{bg.indexed}"
        
        value = cell.value
        if value is not None or fill_type is not None:
            col_letter = get_column_letter(col)
            print(f"  {col_letter}{row}: value={repr(value)}, fill_type={fill_type}, fg={fg_color}, bg={bg_color}")
