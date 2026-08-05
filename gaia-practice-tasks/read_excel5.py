import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx', data_only=False)
ws = wb.active

# Build a grid of colors
color_map = {}
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        color = 'W'
        if fill.patternType:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ('00000000', '0'):
                if '0099FF' in str(rgb):
                    color = 'B'  # BLUE - wall
                elif '92D050' in str(rgb):
                    color = 'G'  # GREEN
                elif 'F478A7' in str(rgb):
                    color = 'P'  # PINK
                elif 'FFFF00' in str(rgb):
                    color = 'Y'  # YELLOW
        val = cell.value
        if val == 'START':
            color = 'S'
        elif val == 'END':
            color = 'E'
        color_map[(row_idx, col_idx)] = color

# Print visual grid
print("Visual map (S=START, E=END, B=BLUE/wall, G=GREEN, P=PINK, Y=YELLOW, W=WHITE/no fill):")
print("   ", end="")
for c in range(1, 10):
    print(f" Col{c}", end="")
print()
for r in range(1, 21):
    print(f"Row{r:>2}: ", end="")
    for c in range(1, 10):
        print(f"  {color_map[(r,c)]} ", end="")
    print()
