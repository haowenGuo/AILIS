import openpyxl

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

# Check START (A1) and END (I20) cells more carefully
for coord in ['A1', 'I20']:
    cell = ws[coord]
    fill = cell.fill
    print(f"\n=== {coord} (value={cell.value}) ===")
    print(f"  fill.patternType: {fill.patternType}")
    fg = fill.fgColor
    print(f"  fgColor.type: {fg.type}")
    print(f"  fgColor.rgb: {fg.rgb}")
    print(f"  fgColor.theme: {fg.theme}")
    print(f"  fgColor.tint: {fg.tint}")
    print(f"  fgColor.indexed: {fg.indexed}")
    bg = fill.bgColor
    print(f"  bgColor.type: {bg.type}")
    print(f"  bgColor.rgb: {bg.rgb}")
    print(f"  bgColor.theme: {bg.theme}")
    print(f"  bgColor.tint: {bg.tint}")
    print(f"  font color: {cell.font.color}")
    if cell.font.color:
        fc = cell.font.color
        print(f"  font.color.type: {fc.type}")
        print(f"  font.color.rgb: {fc.rgb}")
        print(f"  font.color.theme: {fc.theme}")

# Also check if there are any merged cells
print(f"\nMerged cells: {ws.merged_cells.ranges}")

# Let's also check all cells for theme colors
print("\n=== All cells with theme colors ===")
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        fill = cell.fill
        fg = fill.fgColor
        if fg.type == 'theme':
            print(f"  {cell.coordinate}: theme={fg.theme}, tint={fg.tint}, value={cell.value}")
