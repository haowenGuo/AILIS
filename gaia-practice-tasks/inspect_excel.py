import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Find START and END cells
    start_cell = None
    end_cell = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'START' in cell.value.upper():
                    start_cell = cell
                if 'END' in cell.value.upper():
                    end_cell = cell
    
    if start_cell:
        print(f"START cell: {start_cell.coordinate} = '{start_cell.value}', fill: {start_cell.fill.fgColor.rgb if start_cell.fill.fgColor else 'none'}")
    if end_cell:
        print(f"END cell: {end_cell.coordinate} = '{end_cell.value}', fill: {end_cell.fill.fgColor.rgb if end_cell.fill.fgColor else 'none'}")
    
    # Print all cells with their values and fill colors
    print("\n--- Cell grid (value | fill_color) ---")
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value if cell.value is not None else ''
            fill = cell.fill
            fg = fill.fgColor
            if fg and fg.rgb and fg.rgb != '00000000':
                color_str = str(fg.rgb)
            else:
                color_str = 'none'
            row_data.append(f"{get_column_letter(col_idx)}{row_idx}:{val}|{color_str}")
        print(" | ".join(row_data))
