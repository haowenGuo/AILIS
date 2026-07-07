import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb['Sheet1']

# Get START and END cell fills properly
for coord in ['A1', 'I20']:
    cell = ws[coord]
    fill = cell.fill
    print(f"Cell {coord} value='{cell.value}'")
    print(f"  fill.patternType: {fill.patternType}")
    fg = fill.fgColor
    print(f"  fgColor.type: {fg.type}")
    if fg.type == 'rgb':
        print(f"  fgColor.rgb: {fg.rgb}")
    elif fg.type == 'theme':
        print(f"  fgColor.theme: {fg.theme}, tint: {fg.tint}")
    elif fg.type == 'indexed':
        print(f"  fgColor.indexed: {fg.indexed}")
    bg = fill.bgColor
    print(f"  bgColor.type: {bg.type}")
    if bg.type == 'rgb':
        print(f"  bgColor.rgb: {bg.rgb}")

# Build a clean color map
print("\n=== Color Map (rows 1-20, cols A-I) ===")
color_map = {}
for row_idx in range(1, 21):
    for col_idx in range(1, 10):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        fg = fill.fgColor
        if fg.type == 'rgb':
            c = str(fg.rgb)
        elif fg.type == 'theme':
            # theme colors - need to resolve
            c = f"theme:{fg.theme}:{fg.tint}"
        elif fg.type == 'indexed':
            c = f"indexed:{fg.indexed}"
        else:
            c = 'none'
        color_map[(row_idx, col_idx)] = c

# Print as a grid
# Legend: B=blue(obstacle), G=green, P=pink, Y=yellow, S=start, E=end
col_letters = [get_column_letter(c) for c in range(1,10)]
print("    " + " ".join(f"{c:^6}" for c in col_letters))
for r in range(1, 21):
    row_str = f"{r:2d}: "
    for c in range(1, 10):
        val = ws.cell(row=r, column=c).value
        color = color_map[(r,c)]
        if val and 'START' in str(val).upper():
            ch = 'S'
        elif val and 'END' in str(val).upper():
            ch = 'E'
        elif color == 'FF0099FF':
            ch = 'B'
        elif color == 'FF92D050':
            ch = 'G'
        elif color == 'FFF478A7':
            ch = 'P'
        elif color == 'FFFFFF00':
            ch = 'Y'
        else:
            ch = '?'
        row_str += f"  {ch}({color[2:8] if color.startswith('FF') else color})"
    print(row_str)
