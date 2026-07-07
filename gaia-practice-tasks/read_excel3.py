import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx', data_only=False)
ws = wb.active

# Build a grid
grid = {}
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        color = 'NONE'
        if fill.patternType:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ('00000000', '0'):
                color = rgb
        
        val = cell.value
        grid[(row_idx, col_idx)] = (val, color)

# Print the grid with values
print("Grid with values and colors:")
print("     ", end="")
for col in range(1, ws.max_column + 1):
    print(f"Col{col:<5}", end=" ")
print()
for row_idx in range(1, ws.max_row + 1):
    print(f"Row{row_idx:<2}: ", end="")
    for col_idx in range(1, ws.max_column + 1):
        val, color = grid[(row_idx, col_idx)]
        val_str = str(val) if val is not None else ""
        if '0099FF' in str(color):
            val_str = f"BLUE({val_str})"
        print(f"{val_str:<8}", end=" ")
    print()

print()
# Find START and END  
for (r,c), (val, color) in grid.items():
    if val == 'START':
        print(f"START at Row {r}, Col {c} (A{chr(64+c) if c<=26 else ''})")
    if val == 'END':
        print(f"END at Row {r}, Col {c}")
