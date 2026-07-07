import openpyxl

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

# Check all cells in column A carefully
print("Column A detailed:")
for r in range(1, 21):
    cell = ws.cell(row=r, column=1)
    fill = cell.fill
    fg = fill.fgColor
    print(f"  A{r}: value={repr(cell.value)}, fill_type={fill.fill_type}, fg_type={fg.type}, fg_rgb={fg.rgb if fg.type=='rgb' else 'N/A'}, fg_theme={fg.theme if fg.type=='theme' else 'N/A'}, fg_indexed={fg.indexed if fg.type=='indexed' else 'N/A'}")

print()
# Check row 3 carefully
print("Row 3 detailed:")
for c in range(1, 10):
    cell = ws.cell(row=3, column=c)
    fill = cell.fill
    fg = fill.fgColor
    col_letter = chr(ord('A') + c - 1)
    print(f"  {col_letter}3: value={repr(cell.value)}, fill_type={fill.fill_type}, fg_type={fg.type}, fg_rgb={fg.rgb if fg.type=='rgb' else 'N/A'}")

print()
# Check row 4 carefully
print("Row 4 detailed:")
for c in range(1, 10):
    cell = ws.cell(row=4, column=c)
    fill = cell.fill
    fg = fill.fgColor
    col_letter = chr(ord('A') + c - 1)
    print(f"  {col_letter}4: value={repr(cell.value)}, fill_type={fill.fill_type}, fg_type={fg.type}, fg_rgb={fg.rgb if fg.type=='rgb' else 'N/A'}")

# Check for merged cells
print(f"\nMerged cells: {ws.merged_cells.ranges}")

# Check if there are other sheets
print(f"\nSheet names: {wb.sheetnames}")
