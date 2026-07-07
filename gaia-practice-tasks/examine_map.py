import openpyxl

wb = openpyxl.load_workbook("F:\\AILIS_self_evolution_runtime\\gaia-practice-tasks\\task2-excel-map.xlsx")
ws = wb.active
print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Print all cells with their values, coordinates, and fill colors
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        value = cell.value
        # Only print if cell has value or fill
        if value is not None or (fill and fill.patternType and fill.patternType != 'none'):
            fg_rgb = fill.fgColor.rgb if fill.fgColor else 'none'
            bg_rgb = fill.bgColor.rgb if fill.bgColor else 'none'
            pat = fill.patternType if fill.patternType else 'none'
            print(f"  {cell.coordinate}: value={repr(value)}, pattern={pat}, fgColor={fg_rgb}, bgColor={bg_rgb}")
