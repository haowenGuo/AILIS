import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

# Color mapping for display
color_names = {
    'FF0099FF': 'B',  # Blue - blocked
    'FF92D050': 'G',  # Green
    'FFF478A7': 'P',  # Pink
    'FFFFFF00': 'Y',  # Yellow
}

# Build grid
grid = []
color_grid = []
for r in range(1, ws.max_row + 1):
    row = []
    crow = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        fill = cell.fill
        fg = fill.fgColor
        
        # Get RGB
        rgb = None
        if fg.type == 'rgb':
            rgb = fg.rgb
        elif fg.type == 'theme':
            # Theme color - need to resolve
            rgb = f'theme:{fg.theme}:{fg.tint}'
        
        if val == 'START':
            row.append('ST')
            crow.append('START')
        elif val == 'END':
            row.append('EN')
            crow.append('END')
        elif rgb and rgb in color_names:
            row.append(color_names[rgb])
            crow.append(rgb)
        elif rgb:
            row.append('?')
            crow.append(rgb)
        else:
            row.append('.')
            crow.append(None)
    grid.append(row)
    color_grid.append(crow)

# Print grid
print("Map (B=Blue/blocked, G=Green, P=Pink, Y=Yellow, ST=Start, EN=End)")
print("    A  B  C  D  E  F  G  H  I")
for i, row in enumerate(grid):
    print(f"{i+1:2d}: {' '.join(f'{cell:>2}' for cell in row)}")

print("\n\nDetailed color grid (hex codes):")
print("    A        B        C        D        E        F        G        H        I")
for i, crow in enumerate(color_grid):
    print(f"{i+1:2d}: {' '.join(f'{str(c):>8}' for c in crow)}")
