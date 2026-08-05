import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
print("Sheet names:", wb.sheetnames)

ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Print all cells with their values and fill colors
print("\n=== All cells with values or non-default fills ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        fill = cell.fill
        fg = fill.fgColor
        bg = fill.bgColor
        fg_rgb = fg.rgb if fg and fg.rgb else None
        bg_rgb = bg.rgb if bg and bg.rgb else None
        fill_type = fill.patternType
        val = cell.value
        if val is not None or (fill_type and fill_type != 'none'):
            coord = cell.coordinate
            print(f"{coord}: value={repr(val)}, fill_type={fill_type}, fg_rgb={fg_rgb}, bg_rgb={bg_rgb}")
