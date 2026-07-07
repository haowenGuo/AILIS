import openpyxl
from openpyxl.styles import PatternFill

filepath = r"F:\AIGril\eval-results\engineering\gaia-official\files\65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx"

wb = openpyxl.load_workbook(filepath)
print("Sheets:", wb.sheetnames)

ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Find START and END cells, and collect all cell values and colors
start_cell = None
end_cell = None
cells_info = []

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        fill = cell.fill
        fg_color = None
        if fill and fill.fgColor:
            fg_color = fill.fgColor.rgb if fill.fgColor.rgb else None
        
        if val and "START" in str(val).upper():
            start_cell = (row, col, val, fg_color)
            print(f"START found at row={row}, col={col}, value='{val}', color={fg_color}")
        if val and "END" in str(val).upper():
            end_cell = (row, col, val, fg_color)
            print(f"END found at row={row}, col={col}, value='{val}', color={fg_color}")

print(f"\nStart cell: {start_cell}")
print(f"End cell: {end_cell}")

# Print a grid view with values
print("\n--- Grid Values ---")
for row in range(1, ws.max_row + 1):
    row_vals = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value) if cell.value else ""
        row_vals.append(f"{val:8s}")
    print(" | ".join(row_vals))

# Print a grid view with colors
print("\n--- Grid Colors (fgColor rgb) ---")
for row in range(1, ws.max_row + 1):
    row_colors = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        color = ""
        if fill and fill.fgColor and fill.fgColor.rgb:
            color = fill.fgColor.rgb
        row_colors.append(f"{color:10s}")
    print(" | ".join(row_colors))
