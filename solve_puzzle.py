import openpyxl
from openpyxl.styles import PatternFill
import os

# Try both possible file paths
paths = [
    r'F:\AILIS_self_evolution_runtime\map_puzzle.xlsx',
    r'F:\AIGril\eval-results\engineering\gaia-official\files\65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx'
]
file_path = None
for p in paths:
    if os.path.exists(p):
        file_path = p
        print(f'Found file: {p}')
        break

if not file_path:
    print('File not found in either location')
    exit(1)

wb = openpyxl.load_workbook(file_path)
print(f'Sheets: {wb.sheetnames}')
ws = wb.active
print(f'Active sheet: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()

# Find START and END cells, and collect all cell data
start_pos = None
end_pos = None
grid = {}  # (row, col) -> {value, fill_color, font_color}

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        fill = cell.fill
        
        fill_color = None
        if fill and fill.fgColor:
            if fill.fgColor.type == 'rgb':
                fill_color = fill.fgColor.rgb
            elif fill.fgColor.type == 'indexed':
                fill_color = f'indexed:{fill.fgColor.indexed}'
            elif fill.fgColor.type == 'theme':
                fill_color = f'theme:{fill.fgColor.theme}'
        
        # Skip cells with no fill and no value
        if value is None and (fill_color is None or fill.patternType is None):
            continue
        
        grid[(row, col)] = {
            'value': value,
            'fill_color': fill_color,
            'pattern': fill.patternType if fill else None
        }
        
        if value and str(value).strip().upper() == 'START':
            start_pos = (row, col)
            print(f'START found at row={row}, col={col}, fill={fill_color}')
        if value and str(value).strip().upper() == 'END':
            end_pos = (row, col)
            print(f'END found at row={row}, col={col}, fill={fill_color}')

print(f'\nTotal cells with content/fill: {len(grid)}')
print(f'Start: {start_pos}')
print(f'End: {end_pos}')

# Print all unique fill colors
colors = set()
for pos, data in grid.items():
    if data['fill_color']:
        colors.add(data['fill_color'])
print(f'\nUnique fill colors: {sorted(colors)}')

# Print a sample of cells (first 30)
print('\nSample cells:')
for i, (pos, data) in enumerate(sorted(grid.items())):
    if i >= 30:
        break
    print(f'  ({pos[0]},{pos[1]}): value={data["value"]}, fill={data["fill_color"]}, pattern={data["pattern"]}')
