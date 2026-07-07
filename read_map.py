import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(r'F:\AIGril\eval-results\engineering\gaia-official\files\65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active
print('Sheet name:', ws.title)
print('Dimensions:', ws.dimensions)
print('Max row:', ws.max_row, 'Max col:', ws.max_column)
print()

# Find START and END cells
start = None
end = None
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value).strip() if cell.value else ''
        if 'START' in val.upper():
            start = (row, col, val)
        if 'END' in val.upper():
            end = (row, col, val)

print('START:', start)
print('END:', end)
print()

# Print all cell values and colors for the grid
print('=== Grid values and colors ===')
for row in range(1, ws.max_row + 1):
    row_vals = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        fill = cell.fill
        color_hex = 'FFFFFF'
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            # rgb might be an RGB object or a string
            rgb_str = str(rgb)
            if len(rgb_str) == 8:
                color_hex = rgb_str[2:]  # Skip alpha
            elif len(rgb_str) == 6:
                color_hex = rgb_str
        row_vals.append(f'{val if val else "":^6}|{color_hex}')
    print(f'Row {row:2d}: ' + ' '.join(row_vals))
