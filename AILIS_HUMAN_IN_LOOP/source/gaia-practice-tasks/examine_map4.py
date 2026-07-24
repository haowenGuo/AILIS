import openpyxl

wb = openpyxl.load_workbook("F:\\AILIS_self_evolution_runtime\\gaia-practice-tasks\\task2-excel-map.xlsx")
ws = wb.active

lines = []

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        val = cell.value
        
        fg_rgb = "N/A"
        pat = "N/A"
        try:
            if fill and fill.fgColor:
                fg_rgb = fill.fgColor.rgb
        except Exception as e:
            fg_rgb = f"ERR:{e}"
        try:
            if fill:
                pat = fill.patternType
        except Exception as e:
            pat = f"ERR:{e}"
        
        if val is not None or pat not in [None, 'none', 'N/A']:
            lines.append(f"{cell.coordinate}: val={repr(val)}, pat={pat}, fg={fg_rgb}")

# Write to file
with open("F:\\AILIS_self_evolution_runtime\\gaia-practice-tasks\\map_output.txt", "w") as f:
    for line in lines:
        f.write(line + "\n")

print(f"Written {len(lines)} lines to output file")

# Also print just START and END cells specifically
for cell_ref in ['A1', 'I20']:
    cell = ws[cell_ref]
    fill = cell.fill
    val = cell.value
    print(f"\n{cell_ref}: value={repr(val)}")
    print(f"  fill.patternType = {fill.patternType}")
    print(f"  fill.fgColor type = {type(fill.fgColor)}")
    print(f"  fill.fgColor.rgb = {fill.fgColor.rgb if fill.fgColor else 'None'}")
    print(f"  fill.fgColor.theme = {fill.fgColor.theme if fill.fgColor else 'None'}")
    print(f"  fill.fgColor.tint = {fill.fgColor.tint if fill.fgColor else 'None'}")
    print(f"  fill.fgColor.indexed = {fill.fgColor.indexed if fill.fgColor else 'None'}")
    print(f"  fill.fgColor.type = {fill.fgColor.type if fill.fgColor else 'None'}")
    print(f"  fill.bgColor = {fill.bgColor.rgb if fill.bgColor else 'None'}")
