import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx', data_only=False)
ws = wb.active

# Build a grid
grid = {}
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        color = 'NONE'
        if fill.patternType:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ('00000000', '0'):
                color = rgb
        val = cell.value
        grid[(row_idx, col_idx)] = (val, color)

print("Non-blue cells (excluding '0099FF'):")
for (r,c), (val, color) in sorted(grid.items()):
    if '0099FF' not in str(color):
        print(f"  Cell ({r},{c}) [col {c}]: val={val!r}, color={color}")

print()
print("\nAll cells with non-NONE color, ordered:")
for (r,c), (val, color) in sorted(grid.items()):
    if color != 'NONE' and '0099FF' not in str(color):
        print(f"  ({r},{c}): val={val!r}, color={color}")
