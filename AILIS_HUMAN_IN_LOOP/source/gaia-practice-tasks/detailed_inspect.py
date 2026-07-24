import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb['Sheet1']

# Detailed inspection of every cell - check fill pattern type, fgColor, and also check for borders, font colors, etc.
print("=== Detailed cell inspection ===")
for row_idx in range(1, 21):
    for col_idx in range(1, 10):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        fg = fill.fgColor
        bg = fill.bgColor
        
        # Get color info
        color_info = []
        if fg.type == 'rgb':
            color_info.append(f"fg={fg.rgb}")
        elif fg.type == 'theme':
            color_info.append(f"fg=theme{fg.theme}:t{fg.tint}")
        elif fg.type == 'indexed':
            color_info.append(f"fg=idx{fg.indexed}")
        
        if bg.type == 'rgb':
            color_info.append(f"bg={bg.rgb}")
        elif bg.type == 'theme':
            color_info.append(f"bg=theme{bg.theme}:t{bg.tint}")
        elif bg.type == 'indexed':
            color_info.append(f"bg=idx{bg.indexed}")
            
        val = cell.value
        if val is not None:
            color_info.append(f"val='{val}'")
        
        coord = f"{get_column_letter(col_idx)}{row_idx}"
        print(f"{coord}: pattern={fill.patternType}, {', '.join(color_info)}")

# Check for merged cells
print(f"\nMerged cells: {ws.merged_cells.ranges}")

# Check column widths and row heights
print("\nColumn widths:")
for col_letter in ['A','B','C','D','E','F','G','H','I']:
    dim = ws.column_dimensions.get(col_letter)
    if dim:
        print(f"  {col_letter}: width={dim.width}")
