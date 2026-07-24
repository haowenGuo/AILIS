import openpyxl
from openpyxl.utils import get_column_letter
from collections import deque

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb['Sheet1']

BLUE = 'FF0099FF'
ROWS, COLS = 20, 9

def get_color(r, c):
    cell = ws.cell(row=r, column=c)
    fg = cell.fill.fgColor
    if fg.type == 'rgb':
        return str(fg.rgb)
    elif fg.type == 'theme':
        return f'theme:{fg.theme}:{fg.tint}'
    return 'unknown'

def is_blue(r, c):
    return get_color(r, c) == BLUE

def is_walkable(r, c):
    if r < 1 or r > ROWS or c < 1 or c > COLS:
        return False
    return not is_blue(r, c)

def coord(r, c):
    return f"{get_column_letter(c)}{r}"

# Print the grid clearly
print("=== Grid (B=blue/wall, G=green, P=pink, Y=yellow, S=start, E=end) ===")
for r in range(1, ROWS+1):
    line = f"{r:2d}: "
    for c in range(1, COLS+1):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val and 'START' in str(val).upper():
            line += "S "
        elif val and 'END' in str(val).upper():
            line += "E "
        elif is_blue(r, c):
            line += "B "
        else:
            color = get_color(r, c)
            if color == 'FF92D050':
                line += "G "
            elif color == 'FFF478A7':
                line += "P "
            elif color == 'FFFFFF00':
                line += "Y "
            else:
                line += "? "
    print(line)

start = (1, 1)
end = (20, 9)
directions = [(-1,0,'up'), (1,0,'down'), (0,-1,'left'), (0,1,'right')]

print("\n=== Moves from START (A1) ===")
for dr, dc, dname in directions:
    r1, c1 = start[0]+dr, start[1]+dc
    r2, c2 = start[0]+dr*2, start[1]+dc*2
    dest_in_bounds = 1 <= r2 <= ROWS and 1 <= c2 <= COLS
    inter_in_bounds = 1 <= r1 <= ROWS and 1 <= c1 <= COLS
    dest_walkable = dest_in_bounds and is_walkable(r2, c2)
    inter_walkable = inter_in_bounds and is_walkable(r1, c1)
    print(f"  {dname}: inter={coord(r1,c1) if inter_in_bounds else 'OOB'} walkable={inter_walkable}, dest={coord(r2,c2) if dest_in_bounds else 'OOB'} walkable={dest_walkable}")

# BFS with strict rule (both cells walkable), no backward direction
def get_moves_strict(pos, last_dir=None):
    r, c = pos
    moves = []
    for dr, dc, dname in directions:
        if last_dir and (dr, dc) == (-last_dir[0], -last_dir[1]):
            continue
        r1, c1 = r+dr, c+dc
        r2, c2 = r+dr*2, c+dc*2
        if 1 <= r2 <= ROWS and 1 <= c2 <= COLS:
            if is_walkable(r1, c1) and is_walkable(r2, c2):
                moves.append(((r2,c2), dname, (dr,dc)))
    return moves

def get_moves_dest_only(pos, last_dir=None):
    r, c = pos
    moves = []
    for dr, dc, dname in directions:
        if last_dir and (dr, dc) == (-last_dir[0], -last_dir[1]):
            continue
        r2, c2 = r+dr*2, c+dc*2
        if 1 <= r2 <= ROWS and 1 <= c2 <= COLS:
            if is_walkable(r2, c2):
                moves.append(((r2,c2), dname, (dr,dc)))
    return moves

def bfs(start, end, move_func):
    queue = deque()
    queue.append((start, None, [start]))
    visited = set()
    visited.add((start, None))
    
    while queue:
        pos, last_dir, path = queue.popleft()
        if pos == end:
            return path
        for next_pos, dname, dir_vec in move_func(pos, last_dir=last_dir):
            state = (next_pos, dir_vec)
            if state not in visited:
                visited.add(state)
                queue.append((next_pos, dir_vec, path + [next_pos]))
    return None

path_strict = bfs(start, end, get_moves_strict)
path_dest = bfs(start, end, get_moves_dest_only)

print(f"\nStrict path (both cells walkable, no immediate backward): {path_strict}")
if path_strict:
    print(f"  Length: {len(path_strict)-1} moves")
    print(f"  Path: {' -> '.join(coord(r,c) for r,c in path_strict)}")

print(f"\nDest-only path (jump over blue, no immediate backward): {path_dest}")
if path_dest:
    print(f"  Length: {len(path_dest)-1} moves")
    print(f"  Path: {' -> '.join(coord(r,c) for r,c in path_dest)}")

# BFS with no revisiting cells at all
def bfs_no_revisit(start, end, max_paths=10):
    queue = deque()
    queue.append((start, None, frozenset([start]), [start]))
    found_paths = []
    
    while queue and len(found_paths) < max_paths:
        pos, last_dir, visited_set, path = queue.popleft()
        if pos == end:
            found_paths.append(path)
            continue
        for next_pos, dname, dir_vec in get_moves_strict(pos, last_dir=last_dir):
            if next_pos not in visited_set:
                new_visited = visited_set | {next_pos}
                queue.append((next_pos, dir_vec, new_visited, path + [next_pos]))
    return found_paths

print("\n=== No-revisit paths (strict, no cell revisited) ===")
paths_nr = bfs_no_revisit(start, end, max_paths=5)
print(f"Found {len(paths_nr)} paths")
for i, p in enumerate(paths_nr):
    print(f"  Path {i+1} ({len(p)-1} moves): {' -> '.join(coord(r,c) for r,c in p)}")
