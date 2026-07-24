import openpyxl

wb = openpyxl.load_workbook("F:\\AILIS_self_evolution_runtime\\gaia-practice-tasks\\task2-excel-map.xlsx")
ws = wb.active

# Build a complete grid representation
print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
print()

# First, let's find START and END positions
start_pos = None
end_pos = None
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value == 'START':
            start_pos = (row, col)
        if cell.value == 'END':
            end_pos = (row, col)

print(f"START at: row={start_pos[0]}, col={start_pos[1]}")
print(f"END at: row={end_pos[0]}, col={end_pos[1]}")
print()

# Build a color grid
# Color codes observed:
# FF0099FF - blue (avoid)
# FF92D050 - green
# FFF478A7 - pink
# FFFFFF00 - yellow

# Let's determine each cell's color
color_map = {}
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        color = None
        if cell.value == 'START' or cell.value == 'END':
            # Need to find the fill color for START/END cells
            pass
        if fill and fill.fgColor and fill.fgColor.rgb:
            try:
                color = str(fill.fgColor.rgb)
            except:
                pass
        color_map[(row, col)] = color

# Print the grid
for row in range(1, ws.max_row + 1):
    row_str = f"Row {row:2d}: "
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        color = color_map.get((row, col))
        val = cell.value
        
        if val == 'START':
            row_str += "STRT "
        elif val == 'END':
            row_str += "END  "
        elif color and '0099FF' in color:
            row_str += "BLU  "  # blue - avoid
        elif color and '92D050' in color:
            row_str += "GRN  "  # green
        elif color and 'F478A7' in color:
            row_str += "PNK  "  # pink
        elif color and 'FFFF00' in color:
            row_str += "YEL  "  # yellow
        else:
            row_str += "?    "
    print(row_str)

print()

# Print detailed color map
print("Detailed cell info:")
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        val = cell.value
        if val == 'START' or val == 'END':
            print(f"  {cell.coordinate}: value={repr(val)}")
        else:
            try:
                fg = fill.fgColor.rgb if fill and fill.fgColor else 'none'
            except:
                fg = 'error'
            try:
                pat = fill.patternType if fill and fill.patternType else 'none'
            except:
                pat = 'error'
            if pat != 'none' or val is not None:
                print(f"  {cell.coordinate}: value={repr(val)}, pattern={pat}, fgColor={fg}")
