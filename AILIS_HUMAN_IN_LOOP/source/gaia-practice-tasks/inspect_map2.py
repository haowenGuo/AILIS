import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('task2-excel-map.xlsx')
ws = wb.active

# Let's check fill pattern types and proper fill colors
print("Detailed fill info for all non-empty cells:")
print("-" * 100)

for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        fill = cell.fill
        fg = fill.fgColor
        bg = fill.bgColor
        
        fg_rgb = None
        bg_rgb = None
        if fg.type == 'rgb':
            fg_rgb = fg.rgb
        if bg.type == 'rgb':
            bg_rgb = bg.rgb
        
        # Only print cells with non-default fills
        if fill.patternType is not None or cell.value is not None:
            coord = f"{get_column_letter(c)}{r}"
            print(f"{coord}: val='{cell.value}', pattern={fill.patternType}, "
                  f"fg_type={fg.type}, fg_rgb={fg_rgb}, fg_theme={fg.theme if fg.type=='theme' else '-'}, "
                  f"fg_indexed={fg.indexed if fg.type=='indexed' else '-'}, "
                  f"bg_type={bg.type}, bg_rgb={bg_rgb}, bg_indexed={bg.indexed if bg.type=='indexed' else '-'}")
