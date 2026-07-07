import openpyxl

wb = openpyxl.load_workbook(r'F:\AILIS_self_evolution_runtime\gaia-practice-tasks\task2-excel-map.xlsx')
ws = wb.active

# Print exact grid with colors for every cell
# Rows 1-20, Cols A-I (1-9)
print("Complete grid with hex colors (RRGGBB portion):")
print("     A       B       C       D       E       F       G       H       I")
for row in range(1, ws.max_row + 1):
    line = f"{row:2d}: "
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        fill = cell.fill
        fg = fill.fgColor
        fg_color = "      "
        if fg and fg.type == 'rgb':
            # ARGB format, take last 6 chars for RRGGBB
            rgb = fg.rgb
            if len(rgb) == 8:
                fg_color = rgb[2:]  # skip alpha
            else:
                fg_color = rgb
        elif fg and fg.type == 'theme':
            fg_color = f"th{fg.theme}"
        
        if val:
            fg_color = val[:6].ljust(6)
        line += f"{fg_color:>7} "
    print(line)
