import openpyxl
from openpyxl.styles import PatternFill

filepath = r"F:\AIGril\eval-results\engineering\gaia-official\files\65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d-65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx"

wb = openpyxl.load_workbook(filepath)
ws = wb.active

print(f"Grid: {ws.max_row} rows x {ws.max_column} cols")

# Build grid with values and colors
grid = []
for row in range(1, ws.max_row + 1):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        fill = cell.fill
        color_hex = None
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            # rgb could be an RGB object or a string
            if hasattr(rgb, 'value'):
                color_hex = str(rgb.value)
            else:
                color_hex = str(rgb)
        row_data.append((val, color_hex))
    grid.append(row_data)

# Find START and END
start = None
end = None
for r in range(len(grid)):
    for c in range(len(grid[r])):
        val = grid[r][c][0]
        if val and "START" in str(val).upper():
            start = (r, c)
            print(f"START at row={r+1}, col={c+1}, value='{val}', color={grid[r][c][1]}")
        if val and "END" in str(val).upper():
            end = (r, c)
            print(f"END at row={r+1}, col={c+1}, value='{val}', color={grid[r][c][1]}")

# Print all unique colors
all_colors = set()
for r in range(len(grid)):
    for c in range(len(grid[r])):
        if grid[r][c][1]:
            all_colors.add(grid[r][c][1])
print(f"\nAll unique colors: {sorted(all_colors)}")

# Identify blue cells
# Blue colors typically have high B component and low R/G
blue_cells = []
for r in range(len(grid)):
    for c in range(len(grid[r])):
        color = grid[r][c][1]
        if color and len(color) >= 6:
            # Handle ARGB format (8 chars) or RGB (6 chars)
            hex_color = color[-6:]  # take last 6 chars (RGB part)
            try:
                r_val = int(hex_color[0:2], 16)
                g_val = int(hex_color[2:4], 16)
                b_val = int(hex_color[4:6], 16)
                # Blue-ish: blue dominant
                if b_val > 150 and b_val > r_val + 50 and b_val > g_val + 50:
                    blue_cells.append((r+1, c+1, color))
            except:
                pass

print(f"\nBlue-ish cells ({len(blue_cells)}):")
for cell in blue_cells:
    print(f"  Row {cell[0]}, Col {cell[1]}: {cell[2]}")

# Print a compact map
print("\n--- Map (S=START, E=END, B=blue, .=other) ---")
for r in range(len(grid)):
    row_str = ""
    for c in range(len(grid[r])):
        val = grid[r][c][0]
        color = grid[r][c][1]
        if val and "START" in str(val).upper():
            row_str += "S"
        elif val and "END" in str(val).upper():
            row_str += "E"
        else:
            # check if blue
            is_blue = False
            if color and len(color) >= 6:
                hex_color = color[-6:]
                try:
                    r_val = int(hex_color[0:2], 16)
                    g_val = int(hex_color[2:4], 16)
                    b_val = int(hex_color[4:6], 16)
                    if b_val > 150 and b_val > r_val + 50 and b_val > g_val + 50:
                        is_blue = True
                except:
                    pass
            if is_blue:
                row_str += "B"
            else:
                row_str += "."
    print(row_str)

# Now solve the path problem
# Rules:
# - Start at START cell
# - Move 2 cells per turn (up, down, left, right)
# - Cannot move fewer than 2 cells
# - Cannot move backward (reverse direction of previous move)
# - Cannot land on blue cells
# - What cell do you land on after 11th turn?

print("\n\n=== Path Finding ===")
from collections import deque

# BFS with state (row, col, last_direction, turn_count)
# Directions: up(-1,0), down(1,0), left(0,-1), right(0,1)
# "backward" means opposite direction of last move

dirs = {
    'up': (-2, 0),
    'down': (2, 0),
    'left': (0, -2),
    'right': (0, 2)
}
opposite = {
    'up': 'down',
    'down': 'up',
    'left': 'right',
    'right': 'left',
    None: None
}

def is_blue(r, c):
    if r < 0 or r >= len(grid) or c < 0 or c >= len(grid[0]):
        return True  # out of bounds treated as blocked
    color = grid[r][c][1]
    if not color:
        return False
    if len(color) >= 6:
        hex_color = color[-6:]
        try:
            r_val = int(hex_color[0:2], 16)
            g_val = int(hex_color[2:4], 16)
            b_val = int(hex_color[4:6], 16)
            if b_val > 150 and b_val > r_val + 50 and b_val > g_val + 50:
                return True
        except:
            pass
    return False

def get_color(r, c):
    color = grid[r][c][1]
    if not color:
        return "FFFFFF"  # default white
    if len(color) >= 6:
        return color[-6:]
    return color

# BFS: we want to find all reachable cells after exactly 11 turns
# But we need to know the "correct" path - probably the one that reaches END?
# Actually, the problem says "move toward the END cell" - so there's a specific path.
# Let me think... "you start on START and move toward END" - implies a unique path.
# With the constraints (2 cells per turn, no backward, avoid blue), there might be only one valid path.

# Let's do BFS tracking (r, c, last_dir, turns) and find all cells reachable in exactly 11 turns
# But also check if END is reachable and in how many turns

# Actually, let me re-read: "On the eleventh turn, what is the 6-digit hex code of the color of the cell where you land after moving?"
# This implies there's a specific path. Let me see if there's a unique path from START to END following these rules.

# Let's do DFS to find path to END
start_r, start_c = start
end_r, end_c = end

print(f"Start: ({start_r}, {start_c}), End: ({end_r}, {end_c})")
print(f"Start color: {get_color(start_r, start_c)}")
print(f"End color: {get_color(end_r, end_c)}")

# Find all paths to END with exactly N turns, or find shortest path
# Since the problem asks about the 11th turn, the path likely has 11 turns to reach some cell
# (maybe not END yet, or END is reached on 11th turn)

# Let's do BFS to find reachable states after each turn
# State: (r, c, last_direction)
# We track visited per turn to avoid infinite loops (but we can revisit cells from different directions)

# Actually with "no backward" rule, we can't immediately reverse, but we can loop around.
# Let's find all cells reachable after exactly 11 turns.

# Use BFS level by level
current_states = {(start_r, start_c, None)}  # (r, c, last_dir)

for turn in range(1, 12):  # turns 1 to 11
    next_states = set()
    for (r, c, last_dir) in current_states:
        for dname, (dr, dc) in dirs.items():
            # Can't go backward
            if last_dir and dname == opposite[last_dir]:
                continue
            nr, nc = r + dr, c + dc
            # Check bounds
            if nr < 0 or nr >= len(grid) or nc < 0 or nc >= len(grid[0]):
                continue
            # Check blue
            if is_blue(nr, nc):
                continue
            next_states.add((nr, nc, dname))
    current_states = next_states
    print(f"After turn {turn}: {len(current_states)} reachable states")
    if not current_states:
        print("No more moves possible!")
        break

# After 11 turns, what cells can we be on?
cells_after_11 = set()
for (r, c, d) in current_states:
    cells_after_11.add((r, c, get_color(r, c)))

print(f"\nCells reachable after 11 turns: {len(cells_after_11)}")
for r, c, color in sorted(cells_after_11):
    val = grid[r][c][0]
    print(f"  Row {r+1}, Col {c+1} (color={color}, value={val})")

# Check if END is among them
for r, c, color in cells_after_11:
    if (r, c) == (end_r, end_c):
        print(f"\nEND is reachable after 11 turns! Color: {color}")
        break
else:
    print("\nEND is NOT reachable after 11 turns.")
    # Check when END becomes reachable
    current_states2 = {(start_r, start_c, None)}
    for turn in range(1, 30):
        next_states = set()
        for (r, c, last_dir) in current_states2:
            for dname, (dr, dc) in dirs.items():
                if last_dir and dname == opposite[last_dir]:
                    continue
                nr, nc = r + dr, c + dc
                if nr < 0 or nr >= len(grid) or nc < 0 or nc >= len(grid[0]):
                    continue
                if is_blue(nr, nc):
                    continue
                next_states.add((nr, nc, dname))
        current_states2 = next_states
        for (r, c, d) in current_states2:
            if (r, c) == (end_r, end_c):
                print(f"END first reachable at turn {turn}")
                break
        else:
            continue
        break
